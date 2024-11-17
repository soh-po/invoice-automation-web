import os
import shutil
from glob import glob
from datetime import datetime
import openpyxl as opx
import pandas as pd
from time import time
from django.conf import settings


def create_invoice(invoice_template_path, tmp_dir, uuid_dir, create_dates):
    # sales_data_pathは売上データ, invoice_template_pathは請求書テンプレートファイル, uuid_dirはディレクトリの重複を避けるために使う

    def opx_to_pandas_df():
        df = pd.DataFrame(columns=["日付","購入者","品目", "個数", "値段", "小計"]) # データフレームの列を定義
        # openpyxlでファイルを読み込み、pandasでdfにまとめる
        row = 0 # 行番号を定義。dfでは0から始まる
        for file in sales_data_path: # 売上ーデータ（xlsxファイル）を順に変数fileに代入
            print(f"read_Excel_file: {file}") # 読み込みファイルを表示。最終的にはログファイルに記録したい
            wb_sales = opx.load_workbook(file, data_only=True) # xlsxファイルを数値で読み込む（Excelの計算式から計算結果を数値で書き込む）
            ws_sales = wb_sales.active # アクティブなワークシートを選択
            ws_title = ws_sales.title # シート名の取得（請求月表示に使用）
            max_row = ws_sales.max_row # シートの最終行を取得
            for r in range(4, max_row + 1):
                if ws_sales.cell(r, 1).value is not None: # セルが空でない場合
                    df.loc[row, "日付"] = ws_sales.cell(r, 1).value # 日付をr行のA列から取得
                    df.loc[row, "購入者"] = ws_sales.cell(r, 2).value # 購入者をr行のB列から取得
                    df.loc[row, "品目"] = ws_sales.cell(r, 3).value # 品目をr行のC列から取得
                    df.loc[row, "個数"] = ws_sales.cell(r, 4).value # 個数をr行のD列から取得
                    df.loc[row, "値段"] = ws_sales.cell(r, 5).value # 値段をr行のE列から取得
                    df.loc[row, "小計"] = ws_sales.cell(r, 6).value # 小計をr行のF列から取得
                    row += 1 # 行を+1

            grouped = df.groupby("購入者") # 購入者毎にグループ化
        return grouped, ws_title # グループ化したデータとシート名を返す


    #  データフレームを辞書に変換し、openpyxlでテンプレートファイルに書き込む関数
    def write_to_excel():
        customers = {} # 購入者毎のデータを格納する辞書を作成
        for name, group in grouped:
            customers[name] = group.reset_index(drop=True) # 購入者毎にグループ化した値を辞書に入れる

        for name, customer_df in customers.items():
            wb = opx.load_workbook(invoice_template_path) # 請求書テンプレートファイルを開く
            ws = wb.active # アクティブシートを取得

            for i, row in customer_df.iterrows(): # 購入者毎にループ
                ws["B4"] = row["購入者"] # 請求書テンプレートファイルの"B4"セルに購入者名を入力
                file_name = row["購入者"] # ファイル名
                without_space_file_name = file_name.replace(" ", "") # ファイル名からスペースを削除する
                ws["G3"] = cell_date # 請求書の日付
                ws["C10"] = f"{ws_title}分のご請求" # 件名
                bonding_value = f"{row['品目']}({row['日付'].strftime('%m/%d')})" # 品目と日付を1つのセルに書く(内訳欄)変数を定義
                ws.cell(row=i + 15, column=2, value=bonding_value) # 内訳欄
                ws.cell(row=i + 15, column=5, value=row["個数"]) # 個数欄
                ws.cell(row=i + 15, column=6, value=row["値段"]) # 単価欄
                ws.cell(row=i + 15, column=7, value=row["小計"]) # 金額(税込)欄
                # カーソルの位置などを調整する場合
                # ws.sheet_view.selection[0].activeCell = "A1"
                # ws.sheet_view.selection[0].sqref = "A1"
                # ws.sheet_view.topLeftCell = "A1"
            wb.properties.creator = "Automated Invoice Creation Project" # ファイルの作成者を任意に指定
            wb.properties.lastModifiedBy = "" # ファイルを前回保存者を任意に指定
            wb.save(f"{save_dir}/{without_space_file_name}様.xlsx") # ファイルに保存
            print(f"write_Excel_file: {save_dir}/{without_space_file_name}様.xlsx") # 保存ファイルを表示。最終的にはログファイルに記録したい
            


    # 作成されたファイルをZIPにアーカイブする関数
    def archive_to_zip():
        shutil.make_archive(f"{archived_directory}/{archived_name}", format="zip", root_dir=save_dir)


    # メイン処理
    sales_data_path = glob(f"{tmp_dir}/salesbook/*.xlsx") # フォルダ内のxlsxファイルをリストに入れる
    save_dir = os.path.join(settings.MEDIA_ROOT, "invoices", uuid_dir) # xlsx保存先ディレクトリを指定
    archived_directory = os.path.join(settings.MEDIA_ROOT, "download", uuid_dir) # 圧縮ファイルの保存先ディレクトリを作成
    today = datetime.now() # 日時（当日）を取得

    # 請求書作成日の日付処理
    cell_date_str = create_dates[0] # HTMLフォームから送信された日付を取得
    cell_date_obj = datetime.strptime(cell_date_str, "%Y-%m-%d") # 文字列型をdatetime型に変換
    cell_date = cell_date_obj.strftime(("%Y年%m月%d日")) # datetime型から文字列型に再変換
    # cell_date = today.strftime("%Y年%m月%d日") # プログラム実行時に日付にする場合に使う
    # cell_date = "2024年4月30日" # 任意の日付を指定
    archived_name =  f"invoice_{today.strftime("%Y%m%d_%H%M%S")}" # ZIP圧縮を行う際のファイル名

    # 開始時間取得
    start_time = time()

    os.makedirs(save_dir, exist_ok=True)
    os.makedirs(archived_directory, exist_ok=True)

    grouped, ws_title = opx_to_pandas_df()
    write_to_excel()
    archive_to_zip()


    # 終了時間を取得と実行時間の計算
    end_time = time()
    processing_time = round(end_time - start_time, 2)

    # ダウンロードURLと処理時間を返す
    invoice_url  = os.path.join(settings.MEDIA_URL, "download", uuid_dir, f"{archived_name}.zip")
    return invoice_url, processing_time
